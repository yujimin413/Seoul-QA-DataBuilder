# OpenAI API를 사용하여 질문과 답변 데이터를 정제하는 스크립트

# 주요 기능:
# - `.xlsx` 파일에서 질문과 답변 데이터를 정제하여 더욱 명확한 학습 데이터로 변환
# - 불필요한 정보 제거, 중복된 질문 분리, 특정 패턴 정제
# - OpenAI API를 사용하여 질문을 자동으로 정리 및 분할
# - 정제된 데이터를 각각의 새로운 엑셀 파일로 저장 & 문답모음집_정제.xlsx의 각 시트에 추가 (중복된 데이터는 추가 X)

# 실행 조건:
# - Python 환경에서 실행 (`python 3.refine_qna_sets.py`)
# - OpenAI API 키 설정 필요 (`.env` 파일에 `OPENAI_API_KEY` 저장)
# - `2.classify_qna_set.py` 실행 후 생성된 분류된 질문-답변 데이터가 `3.CLASSIFICATION_RESULTS/` 폴더에 존재해야 함
# - 한글 PDF 저장을 원할 경우 `NanumGothic.ttf` 폰트 필요 (현재 PDF 저장 기능은 비활성화)

# 입력 (Input):
# - `3.CLASSIFICATION_RESULTS/` 폴더 내 **2.classify_qna_set.py 실행 결과** (`{파일명}_{주제명}.xlsx`)
# - 엑셀 파일 컬럼: `데이터셋명`, `질문`, `답변`

# 처리 방식:
# - OpenAI API를 사용하여 질문을 분석하고 필요할 경우 질문을 분리
# - 불필요한 내용 제거 (예: "문의드립니다.", "감사합니다." 등)
# - 원천사가 답변하지 않은 경우 질문을 제거
# - 정제된 데이터를 새로운 `.xlsx` 파일로 저장

# 출력 (Output):
# - `4.REFINED_RESULTS/` 폴더에 정제된 데이터 파일 저장
# - 파일명 형식: `{원본파일명}_정제.xlsx`
# - 엑셀 파일 컬럼: `데이터셋명`, `질문`, `답변`
# - 질문을 분리한 경우, 각 질문이 새로운 행으로 추가됨
# - 정제된 데이터가 없으면 파일을 생성하지 않음

# 사용 방법:
# 1. `2.classify_qna_set.py` 실행 후 `3.CLASSIFICATION_RESULTS/` 폴더에 분류된 `.xlsx` 파일이 생성되었는지 확인
# 2. 터미널에서 `python 3.refine_qna_sets.py` 실행
# 3. `4.REFINED_RESULTS/` 폴더에서 생성된 정제된 엑셀 파일 확인


import os
import openai
import pandas as pd
from dotenv import load_dotenv
from fpdf import FPDF
import re
import openpyxl

# 기본 디렉토리 설정
DEFAULT_INPUT_DIRECTORY = "3.CLASSIFICATION_RESULTS"  # 입력 데이터 폴더
DEFAULT_OUTPUT_DIRECTORY = "4.REFINED_RESULTS"  # 결과 데이터 폴더
REFINED_QNA_FILE = "문답모음집_정제.xlsx" # 정제 결과 데이터를 문답모음집_정제.xlsx의 각 시트에 추가

# OpenAI API Key 설정
# .env 파일 로드 (파일이 없을 경우 종료)
if not os.path.exists(".env"):
    print("[WARNING] .env 파일이 존재하지 않습니다. 프로그램을 종료합니다.")
    exit()
else:
    load_dotenv()
openai.api_key = os.environ.get("OPENAI_API_KEY")

def refine_data(question, answer):
    """
    OpenAI API를 호출하여 질문과 답변 데이터를 정제하는 함수.
    반환값: [(refined_question1, refined_answer1), (refined_question2, refined_answer2), ...]
    """
    import re
    prompt = f"""
    # Task
    아래의 질문과 답변에서 불필요한 내용을 제거하거나 질문을 분리한 학습 데이터를 생성해 주세요.
    질문: "{question}"
    답변: "{answer}"

    # 제약조건
    ## '답변'에서 '원천사에 문의 중입니다.'라는 내용이 있을 경우, 해당 질문과 답변은 제외합니다.
    ## '질문 분리'는 필수 조건이 아닙니다. 완벽하게 매칭 되도록 질문을 분리할 수 없다고 판단 되면, '불필요한 내용 제거'만으로도 충분합니다.
    ## '질문 분리' 민감도는 50%로 설정되어 있습니다. 50% 민감도는 아래 예시를 참고해주세요.

    # 예시
    ## question, answer 예시 1
    "question": "생활이동 데이터에서 이동유형에 관하여 문의드립니다.\n\n서울시 생활이동 데이터에서 이동유형은 출발지와 도착지에 H(야간상주지), H(주간상주지), E(기타)를 부여하여 순서쌍으로 표현된 것으로 이해하고 있습니다.\n\n예를 들면, '11010'에서 '11020'으로의 이동유형이 HW이면, 출발지 '11010'은 해당 이동인구의 야간상주지(H)이고 도착지 '11020'는 주간상주지(W)가 된다고 생각합니다.\n\n그리고, 주야간상주지를 추정하는 방법에 대해서는 메뉴얼과 FAQ의 설명에 따르면 개인별로 주야간상주지가 정해지는 것으로 이해됩니다.\n\n그러면, 개인별로 주야간상주지가 하나 이상 부여받을 수 있는 건가요? 데이터에서 확인해보니 주야간상주지가 유일하게 부여되지 않은 것 같은 느낌을 받았습니다.\n\n예를 들면, 이동유형이 'HH'이면 주야간상주지가 같은 지역이라 출발지와 도착지가 같을 거라고 생각했지만 대부분 다른 지역이였고,\n\n반대로 출발지와 도착지가 같음에도 불구하고 이동유형 순서쌍이 다른 경우(HW, WH, 등)가 존재합니다.\n\n정리하자면, 주야간상주지는 개인별로 유일하게 추정되는지 궁금합니다.\n\n혹시 제가 이동유형이나 주야간상주지의 정의를 잘못 이해하고 있다면, 다시 알려주시면 감사하겠습니다.",
    "answer": "개인별로 주야간상주지는 한 달 동안의 체류 패턴에 따라 여러 개가 될 수 있습니다.\n\n체류지(야간/주간 상주지)는 일정 기간 이상 머문 장소를 기준으로 설정합니다."
    
    ## refined_output 예시 1
    [
    (
        "생활이동 데이터에서 이동유형에 관하여 문의드립니다. 서울시 생활이동 데이터에서 이동유형은 출발지와 도착지에 H(야간상주지), H(주간상주지), E(기타)를 부여하여 순서쌍으로 표현된 것으로 이해하고 있습니다. 개인별로 주야간상주지가 하나 이상 부여될 수 있는 건가요?",
        "개인별로 주야간상주지는 한 달 동안의 체류 패턴에 따라 여러 개가 될 수 있습니다."
    ),
    (
        "이동유형 정의와 주야간상주지의 기준이 무엇인지 궁금합니다. 체류지(야간/주간 상주지)는 어떤 기준으로 설정되나요?",
        "체류지(야간/주간 상주지)는 일정 기간 이상 머문 장소를 기준으로 설정합니다."
    )
    ]


    ## question, answer 예시 2
    question: "아래와 같이 문의했는데 정확한 답변이 아니라서 재문의 드립니다. 이전 문의 시 이미지 첨부드린 걸 확인하시면 정확한 문의내용 확인 가능합니다. 상세히 재작성해 문의드립니다. 1. 일자 동일, 시간대 동일, 행정동코드 동일, 집계구코드 동일 => 4가지가 다 동일한 생활인구가 여러개 데이터 값이 있습니다. 예를들어 24년 11월 28일 14시 행정동코드 11110515 집계구코드 1101070000000 의 데이터는 26개 입니다. 2. 생활인구 csv의 집계구 코드 정렬 시, 1101050000000 가 가장 낮은 숫자이며 생활인구 shp의 집계구 코드 정렬 시, 11010530010001 가 가장 낮은 숫자로 shp에는 csv 상에 있는 집계구코드가 없습니다."
    answer: "jupyter notebook, 파이썬을 활용하여 1번에 대한 확인 결과, 아래 이미지와 같이 집계구코드 1101070000000 의 데이터는 존재하지 않습니다. 또한, 24년 11월 28일14시의 행정동코드가 11110515인 데이터는 아래와 같으며, 중복되는 집계구코드는 없음을 확인하였습니다. 2. 위 집계구 코드는 shp파일과 조인이 가능합니다. 프로그램 수행 내역을 같이 첨부해주시면, 원하시는 답변을 드리는데 더욱 수월할 것 같습니다."

    ## refined_output 예시 2
    [
    (
        "일자 동일, 시간대 동일, 행정동코드 동일, 집계구코드 동일한 생활인구 데이터에서 중복된 데이터 값이 있다고 하셨습니다. 이를 확인할 수 있을까요?",
        "jupyter notebook, 파이썬을 활용하여 확인한 결과, 집계구코드 1101070000000의 데이터는 존재하지 않습니다. 또한, 24년 11월 28일 14시의 행정동코드가 11110515인 데이터는 중복되는 집계구코드가 없음을 확인하였습니다."
    ),
    (
        "생활인구 CSV와 shp 파일 간의 집계구 코드 정렬 차이가 있다고 하셨는데, shp 파일에 CSV 상의 집계구 코드가 없다는 점을 설명해 주실 수 있을까요?",
        "위 집계구 코드는 shp 파일과 조인이 가능합니다. 프로그램 수행 내역을 첨부해주시면, 원하시는 답변을 드리는 데 더욱 수월할 것 같습니다."
    )
    ]


    ## question, answer 예시 3
    "question": "'서울생활인구 대도시권 내외국인' 데이터에서 서울 외 모든 지역의 거주자들 수가 계산된 건지 궁금합니다.\n\n'생활인구순위' 열을 보면 대도시권거주지가 68위까지 나와있는데요.\n\n68위 밑으로는 생략하신 건지 원래 68개밖에 없는 건지 궁금합니다.",
    "answer": "서울시 분석팀 [###]입니다.\n\n#행정구역 코드정보 파일 교체 요청 (@ [###])\n\n문의처리중 행정구역 코드정보 파일내 유입주 부분이 현행화되지 않은 부분을 확인하였습니다.\n\n서울 생활인구> 서울빅데이터 | 서울열린데이터광장 행정구역 코드정보 파일 교체 부탁드립니다!\n\n#문의답변 (@[###])\n\n1. 생활인구순위 부여 방식은 같은 시간대, 같은 행정동 코드별로 정렬하여, 총생활인구수가 클수록 높은 순위가 부여됩니다.\n\n2. 12월 데이터의 경우, '생활인구순위'는 68위 까지 나와있는 것을 확인하였습니다.\n\n현재 열린데이터광장에 있는 https://data.seoul.go.kr/dataVisual/seoul/seoulLivingPopulation.do 내 행정구역 코드정보 파일의 \"유입지코드 시트와\"과 비교 하였을때서울을 제외한 코드는 68개로 확인 됩니다.\n\n3. 금일 문의사항 해결중에 행정구역 코드정보 파일 업데이트 하였습니다.\n\n4. 참고로, 도, 특별시, 광역시, 특별자치도는 모두 수록하지만, 시, 군, 구정보는 서울, 인천, 경기만 확인 됩니다."

    ## refined_output 예시 3
    [
    (
        "'서울생활인구 대도시권 내외국인' 데이터에서 서울 외 모든 지역의 거주자 수가 계산된 건가요? '생활인구순위' 열을 보면 대도시권 거주지가 68위까지 나와 있는데, 68위 밑으로는 생략된 건가요, 아니면 원래 68개밖에 없는 건가요?",
        "서울시 분석팀에서는 서울 외 모든 지역의 거주자 수 계산 여부에 대한 데이터를 명시하지 않았습니다. 또한, 12월 데이터 기준으로 '생활인구순위'는 68위까지 확인되며, 서울을 제외한 행정구역 코드는 68개로 확인되었습니다."
    ),
    (
        "서울시 생활인구 데이터의 행정구역 코드정보 파일이 업데이트되었다고 하셨습니다. 업데이트된 내용은 무엇인가요?",
        "행정구역 코드정보 파일의 유입지 코드 시트와 비교하여 서울을 제외한 코드는 68개로 확인되었고, 해당 파일은 금일 업데이트되었습니다."
    )
    ]

    ## question, answer 예시 4
    "question": "안녕하세요? '생활인구'의 단위는 무엇인가요? 궁금해서 질문 드립니다. 감사합니다."
    "answer": "안녕하세요, [###]주무관입니다. '생활인구'의 단위는 1명 입니다. 감사합니다.   

    ## refined_output 예시 4
    [
    (
        "'생활인구'의 단위는 무엇인가요?",
        "'생활인구'의 단위는 1명입니다."
    )
    ]

    # refined_output 출력 형식 (N은 자연수)
    [(refined_question1, refined_answer1), ..., (refined_questionN, refined_answerN)]

    """
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4-turbo",
            messages=[
                {"role": "system", "content": "You are an assistant specialized in refining and structuring data for AI training."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1000
        )

        # API 응답에서 refined_output 추출
        refined_output = response['choices'][0]['message']['content']
        # print(f"[DEBUG] refined_output:\n{refined_output}\n")

        # 정제된 데이터 파싱
        refined_pairs = []
        try:
            # Regex로 질문/답변 쌍 추출
            matches = re.findall(r'\(\s*"(.*?)"\s*,\s*"(.*?)"\s*\)', refined_output, re.DOTALL)
            refined_pairs = [(q.strip(), a.strip()) for q, a in matches if q.strip() and a.strip()]
        except Exception as regex_error:
            # print(f"[DEBUG] Regex parsing error: {regex_error}")
            # print("[DEBUG] Fallback to line-by-line parsing")
            # Line-by-line fallback parsing
            current_question = None
            current_answer = None
            for line in refined_output.strip().split("\n"):
                line = line.strip()
                if line.startswith("질문:"):
                    if current_question and current_answer:
                        refined_pairs.append((current_question.strip(), current_answer.strip()))
                    current_question = line.replace("질문:", "").strip()
                    current_answer = None
                elif line.startswith("답변:"):
                    current_answer = line.replace("답변:", "").strip()
            # 마지막 쌍 추가
            if current_question and current_answer:
                refined_pairs.append((current_question.strip(), current_answer.strip()))

        # print(f"[DEBUG] Refined pairs: {refined_pairs}")
        return refined_pairs

    except Exception as e:
        print(f"Error: {e}")
        return []

def save_to_pdf(df, filename):
    """
    DataFrame을 PDF 파일로 저장하는 함수 (한글 지원)
    """
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # 한글 폰트 설정 (나눔고딕 또는 바탕체 사용)
    pdf.add_font("NanumGothic", "", "NanumGothic.ttf", uni=True)  # 나눔고딕 폰트 추가
    pdf.set_font("NanumGothic", size=10)  # 한글 폰트 사용

    pdf.cell(200, 10, txt="정제된 질문-답변 데이터", ln=True, align='C')
    pdf.ln(10)

    # 테이블 헤더
    col_width = 60
    pdf.set_font("NanumGothic", style='B', size=10)
    pdf.cell(col_width, 10, "데이터셋명", border=1)
    pdf.cell(col_width, 10, "질문", border=1)
    pdf.cell(col_width, 10, "답변", border=1)
    pdf.ln()

    # 테이블 내용
    pdf.set_font("NanumGothic", size=8)
    for _, row in df.iterrows():
        pdf.cell(col_width, 10, row['데이터셋명'][:10] + "...", border=1)
        pdf.cell(col_width, 10, row['질문'][:30] + "...", border=1)
        pdf.cell(col_width, 10, row['답변'][:30] + "...", border=1)
        pdf.ln()

    pdf.output(filename)
    print(f"[INFO] 정제된 데이터가 PDF로 저장되었습니다: {filename}")


def get_unique_filename(directory, base_filename, extension):
    """
    파일명이 중복될 경우 '(1)', '(2)' 등의 숫자를 추가하여 고유한 파일명을 반환.
    """
    filename = os.path.join(directory, f"{base_filename}.{extension}")
    counter = 1

    while os.path.exists(filename):
        filename = os.path.join(directory, f"{base_filename}({counter}).{extension}")
        counter += 1

    return filename


def process_file(input_file):
    """
    입력된 엑셀 파일을 처리하고 정제된 데이터를 새 파일로 저장하는 함수.
    """

    df = pd.read_excel(input_file)
    if '질문' not in df.columns or '답변' not in df.columns:
        print(f"[ERROR] 파일 '{input_file}'에 '질문' 또는 '답변' 컬럼이 없습니다.")
        return

    refined_data = []

    for index, row in df.iterrows():
        dataset_name = row['데이터셋명']
        question = row['질문']
        answer = row['답변']
        # print(f"[DEBUG] Processing row {index + 1}...")

        # API 호출
        refined_pairs = refine_data(question, answer)

        # 결과 저장
        for refined_question, refined_answer in refined_pairs:
            refined_data.append({"데이터셋명": dataset_name, "질문": refined_question, "답변": refined_answer})

    # 결과를 데이터프레임으로 변환
    refined_df = pd.DataFrame(refined_data)

    # 결과 저장 폴더 생성
    os.makedirs(DEFAULT_OUTPUT_DIRECTORY, exist_ok=True)

    base_filename = os.path.splitext(os.path.basename(input_file))[0]
    output_xlsxfile = get_unique_filename(DEFAULT_OUTPUT_DIRECTORY, f"{base_filename}_정제", "xlsx")
    # output_pdf_file = get_unique_filename(DEFAULT_OUTPUT_DIRECTORY, f"{base_filename}_정제", "pdf")

    # 데이터가 없으면 저장하지 않고 경고 메시지 출력
    if refined_df.empty:
        print(f"[WARNING] '{base_filename}' 파일에서 정제된 데이터가 없습니다. 저장하지 않습니다.")
    else:
        refined_df.to_excel(output_xlsxfile, index=False, engine='openpyxl')
        print(f"[INFO] 정제된 데이터가 '{output_xlsxfile}'에 저장되었습니다.")

    # PDF 저장
    # save_to_pdf(refined_df, output_pdf_file)

    # REFINED_QNA_FILE의 적절한 시트에 추가하는 로직
    sheet_mapping = {
        "생활인구(문의)": r"생활인구\(문의\)",
        "생활이동데이터(문의)": r"생활이동데이터\(문의\)",
        "시민생활데이터(문의)": r"시민생활데이터\(문의\)"
    }

    matched_sheet = None
    for sheet_name, pattern in sheet_mapping.items():
        if re.search(pattern, base_filename):
            matched_sheet = sheet_name
            break

    if matched_sheet:
        refined_qna_path = REFINED_QNA_FILE

        # 기존 파일이 있는지 확인하고 불러오기
        if os.path.exists(refined_qna_path):
            book = openpyxl.load_workbook(refined_qna_path)
        else:
            book = openpyxl.Workbook()

        # 해당 시트가 없으면 생성
        if matched_sheet not in book.sheetnames:
            print(f"[ERROR] {matched_sheet} 시트가 존재하지 않습니다.")
            book.create_sheet(matched_sheet)

        sheet = book[matched_sheet]

        # 기존 데이터 읽기 (데이터 중복 방지)
        existing_data = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # print("[INFO] 중복된 데이터가 존재합니다. 추가하지 않고 스킵합니다.")
            existing_data.add(tuple(row))

        # 새로운 데이터 추가
        with pd.ExcelWriter(refined_qna_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            refined_df.to_excel(writer, sheet_name=matched_sheet, index=False, header=not existing_data)

        print(f"[INFO] 정제된 데이터가 '{REFINED_QNA_FILE}'의 '{matched_sheet}' 시트에 추가되었습니다.")


def main():
    print("=" * 80)
    print("OpenAI API를 사용하여 질문-답변 데이터를 정제하는 프로그램")
    print("=" * 80)

    while True:
        choice = input(f"1. 기본 폴더({DEFAULT_INPUT_DIRECTORY}) 내 모든 파일 처리\n"
                       "2. 특정 폴더 입력\n"
                       "3. 프로그램 종료\n"
                       "선택 (1, 2, 3): ").strip()

        if choice == "1":
            directory = DEFAULT_INPUT_DIRECTORY
            break
        elif choice == "2":
            directory = input("정제할 파일이 있는 폴더명을 입력하세요: ").strip()
            if not directory:
                print("[WARNING] 입력된 폴더명이 없습니다. 다시 입력하세요.")
                continue
            break
        elif choice == "3":
            print("프로그램을 종료합니다.")
            return
        else:
            print("[ERROR] 잘못된 입력입니다. 1, 2 또는 3을 입력하세요.\n")

    if not os.path.exists(directory):
        print(f"[ERROR] 디렉토리가 존재하지 않습니다: {directory}")
        return

    print(f"[INFO] '{directory}' 폴더 내 모든 .xlsx 파일을 처리합니다...")

    xlsx_files = [f for f in os.listdir(directory) if f.endswith(".xlsx")]

    if not xlsx_files:
        print(f"[WARNING] '{directory}' 내에 .xlsx 파일이 없습니다.")
        return

    for file_name in xlsx_files:
        file_path = os.path.join(directory, file_name)
        print(f"[INFO] '{file_name}' 처리 중...")
        process_file(file_path)

    print("[INFO] 모든 파일 처리가 완료되었습니다.")


if __name__ == "__main__":
    main()